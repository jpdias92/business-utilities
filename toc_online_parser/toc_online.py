import pandas as pd
import glob
import re
import os
import argparse
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

def process_files(input_dir, output_file):
    # Find all files in the input directory that match the given pattern
    file_pattern = os.path.join(input_dir, 'bank_statement*.xlsx')
    file_paths = glob.glob(file_pattern)

    if not file_paths:
        print(f"No files found matching the pattern: {file_pattern}")
        return

    # Create an empty DataFrame to store merged data
    merged_data = pd.DataFrame(columns=["Data valor", "Descrição", "Valor", "Descritivo Banco"])

    for file_path in file_paths:
        print(f"Processing file: {file_path}")

        # Read the excel file and skip the first 14 rows
        df = pd.read_excel(file_path, skiprows=14)

        # Filter the relevant columns
        if "Data valor" in df.columns and "Descrição" in df.columns and "Débito" in df.columns and "Crédito" in df.columns:
            # Create a new dataframe with the required columns
            df_filtered = df[["Data valor", "Descrição", "Débito", "Crédito"]]

            # Convert "Data valor" to the desired date format (dd/mm/yyyy)
            df_filtered["Data valor"] = pd.to_datetime(df_filtered["Data valor"], format='%d-%m-%Y')

            # Merge "Débito" and "Crédito" into the "Valor" column
            df_filtered["Valor"] = df_filtered.apply(
                lambda row: row["Débito"] * -1 if pd.notna(row["Débito"]) else row["Crédito"], axis=1
            )

            # Remove the matching pattern from "Descrição" and create "Descritivo Banco"
            df_filtered["Descritivo Banco"] = df_filtered["Descrição"]
            df_filtered["Descrição"] = df_filtered["Descrição"].apply(lambda x: re.sub(r'Transação por cartão de .* emitida por ', '', x))

            # Keep only the required columns
            df_filtered = df_filtered[["Data valor", "Descrição", "Valor", "Descritivo Banco"]]

            # Append to the merged data DataFrame
            merged_data = pd.concat([merged_data, df_filtered], ignore_index=True)

    # Sort the merged data by "Data valor"
    merged_data = merged_data.sort_values(by="Data valor", ascending=True)

    # Save the merged data to the output Excel file
    merged_data.to_excel(output_file, index=False)

    # Open the newly created Excel file to format the "Data valor" column as Date
    wb = load_workbook(output_file)
    ws = wb.active

    # Define a date style to format the "Data valor" column
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
    if "Data valor" in merged_data.columns:
        date_col = merged_data.columns.get_loc("Data valor") + 1  # 1-based index for Excel
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=date_col, max_col=date_col):
            for cell in row:
                cell.style = date_style

    # Save the file with the date formatting applied
    wb.save(output_file)
    print(f"Data successfully merged, sorted, and saved to {output_file} with date formatting applied.")

if __name__ == "__main__":
    # Set up argument parsing
    parser = argparse.ArgumentParser(description="Merge and process TOC Online exported transaction files.")
    
    parser.add_argument("--input", "-i", required=True, help="Directory containing the input Excel files (bank_statement*.xlsx)")
    parser.add_argument("--output", "-o", default="toc_online_transactions.xlsx", help="Output Excel file to save the merged data")

    # Parse the arguments
    args = parser.parse_args()

    # Call the process_files function with the parsed arguments
    process_files(args.input, args.output)
